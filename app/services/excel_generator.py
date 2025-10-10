"""
Excel generation service using openpyxl
"""

import os
import asyncio
from pathlib import Path
from typing import Dict, Any, Optional
import logging

from app.core.config import settings

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Service for generating Excel reports"""
    
    def __init__(self):
        self.temp_dir = Path(settings.temp_dir)
        self.temp_dir.mkdir(exist_ok=True)
    
    async def generate_from_data(
        self,
        data: Dict[str, Any],
        report_id: str,
        options: Optional[Dict[str, Any]] = None
    ) -> str:
        """Generate Excel file from structured data"""
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
        except ImportError:
            raise ImportError("openpyxl is required for Excel generation")
        
        try:
            options = options or {}
            
            # Create output file path
            output_path = self.temp_dir / f"{report_id}.xlsx"
            
            # Create workbook and worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Weather Report"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            row = 1
            
            # Add title
            worksheet.merge_cells(f'A{row}:D{row}')
            title_cell = worksheet[f'A{row}']
            title_cell.value = data.get('title', 'Weather Report')
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            row += 2
            
            # Add location information
            if 'location' in data:
                worksheet[f'A{row}'].value = "Location:"
                worksheet[f'B{row}'].value = data['location']
                worksheet[f'A{row}'].font = Font(bold=True)
                row += 1
            
            if 'coordinates' in data:
                coords = data['coordinates']
                worksheet[f'A{row}'].value = "Coordinates:"
                worksheet[f'B{row}'].value = f"{coords.get('lat', 'N/A')}, {coords.get('lng', 'N/A')}"
                worksheet[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 1
            
            # Add current weather data
            if 'current' in data:
                worksheet[f'A{row}'].value = "Current Weather"
                worksheet[f'A{row}'].font = header_font
                worksheet[f'A{row}'].fill = header_fill
                worksheet[f'A{row}'].border = border
                worksheet[f'B{row}'].border = border
                worksheet[f'C{row}'].border = border
                worksheet[f'D{row}'].border = border
                row += 1
                
                current = data['current']
                weather_items = [
                    ("Temperature", current.get('temperature', 'N/A')),
                    ("Condition", current.get('condition', 'N/A')),
                    ("Wind Speed", current.get('windSpeed', 'N/A')),
                    ("Wind Direction", current.get('windDirection', 'N/A')),
                    ("Humidity", current.get('humidity', 'N/A')),
                    ("Pressure", current.get('pressure', 'N/A'))
                ]
                
                for item, value in weather_items:
                    worksheet[f'A{row}'].value = item
                    worksheet[f'B{row}'].value = value
                    worksheet[f'A{row}'].border = border
                    worksheet[f'B{row}'].border = border
                    worksheet[f'A{row}'].font = Font(bold=True)
                    row += 1
                
                row += 1
            
            # Add forecast data
            if 'forecast' in data and 'periods' in data['forecast']:
                worksheet[f'A{row}'].value = "Weather Forecast"
                worksheet[f'A{row}'].font = header_font
                worksheet[f'A{row}'].fill = header_fill
                worksheet[f'A{row}'].border = border
                worksheet[f'B{row}'].border = border
                worksheet[f'C{row}'].border = border
                worksheet[f'D{row}'].border = border
                row += 1
                
                # Headers
                headers = ["Period", "Temperature", "Condition", "Wind"]
                for col, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                row += 1
                
                # Forecast data
                for period in data['forecast']['periods'][:7]:  # Next 7 periods
                    worksheet[f'A{row}'].value = period.get('name', 'N/A')
                    worksheet[f'B{row}'].value = f"{period.get('temperature', 'N/A')}°{period.get('temperatureUnit', 'F')}"
                    worksheet[f'C{row}'].value = period.get('shortForecast', 'N/A')
                    worksheet[f'D{row}'].value = period.get('windSpeed', 'N/A')
                    
                    for col in range(1, 5):
                        worksheet.cell(row=row, column=col).border = border
                    row += 1
                
                row += 1
            
            # Add historical data if available
            if 'historical' in data:
                worksheet[f'A{row}'].value = "Historical Data Summary"
                worksheet[f'A{row}'].font = header_font
                worksheet[f'A{row}'].fill = header_fill
                worksheet[f'A{row}'].border = border
                worksheet[f'B{row}'].border = border
                worksheet[f'C{row}'].border = border
                worksheet[f'D{row}'].border = border
                row += 1
                
                historical = data['historical']
                hist_items = [
                    ("Average Temperature", historical.get('avgTemperature', 'N/A')),
                    ("Total Precipitation", historical.get('totalPrecipitation', 'N/A')),
                    ("Storm Events", historical.get('stormEvents', 'N/A')),
                    ("Extreme Weather Days", historical.get('extremeWeatherDays', 'N/A'))
                ]
                
                for item, value in hist_items:
                    worksheet[f'A{row}'].value = item
                    worksheet[f'B{row}'].value = value
                    worksheet[f'A{row}'].border = border
                    worksheet[f'B{row}'].border = border
                    worksheet[f'A{row}'].font = Font(bold=True)
                    row += 1
            
            # Add risk assessment if available
            if 'riskAssessment' in data:
                row += 1
                worksheet[f'A{row}'].value = "Risk Assessment"
                worksheet[f'A{row}'].font = header_font
                worksheet[f'A{row}'].fill = header_fill
                worksheet[f'A{row}'].border = border
                worksheet[f'B{row}'].border = border
                worksheet[f'C{row}'].border = border
                worksheet[f'D{row}'].border = border
                row += 1
                
                risk = data['riskAssessment']
                risk_items = [
                    ("Overall Risk Level", risk.get('overallRisk', 'N/A')),
                    ("Weather Risk", risk.get('weatherRisk', 'N/A')),
                    ("Storm Risk", risk.get('stormRisk', 'N/A')),
                    ("Recommendations", risk.get('recommendations', 'N/A'))
                ]
                
                for item, value in risk_items:
                    worksheet[f'A{row}'].value = item
                    worksheet[f'B{row}'].value = value
                    worksheet[f'A{row}'].border = border
                    worksheet[f'B{row}'].border = border
                    worksheet[f'A{row}'].font = Font(bold=True)
                    row += 1
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            workbook.save(str(output_path))
            
            logger.info(f"Excel file generated: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Failed to generate Excel file: {e}")
            raise